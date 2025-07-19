#!/usr/bin/env python3
"""
Spanish Bank Phone Number Extractor - Modern PyQt6 GUI

This is the main GUI entrypoint for the application.
"""

import sys
import os
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QComboBox, QTextEdit, QLineEdit, QPushButton, QLabel, QFileDialog,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QProgressBar, QSplitter, QFrame, QScrollArea,
    QDialog, QDialogButtonBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize
from PyQt6.QtGui import QFont, QPalette, QColor, QIcon, QPixmap

from ..core.extractor import SpanishBankExtractor
from ..core.bank_registry import BankRegistry


class ModernButton(QPushButton):
    """Custom modern button with styling."""
    
    def __init__(self, text: str, primary: bool = False):
        super().__init__(text)
        self.setMinimumHeight(40)
        self.setFont(QFont("Segoe UI", 10))
        
        if primary:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #0078d4;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background-color: #106ebe;
                }
                QPushButton:pressed {
                    background-color: #005a9e;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #f3f2f1;
                    color: #323130;
                    border: 1px solid #d2d0ce;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #edebe9;
                    border-color: #c7c6c4;
                }
                QPushButton:pressed {
                    background-color: #e1dfdd;
                }
                QPushButton:disabled {
                    background-color: #f3f2f1;
                    color: #a19f9d;
                    border-color: #edebe9;
                }
            """)


class FileLoaderThread(QThread):
    """Thread for loading files asynchronously to prevent GUI freezing."""
    
    file_loaded = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)
    
    def __init__(self, file_path: str, file_extension: str):
        super().__init__()
        self.file_path = file_path
        self.file_extension = file_extension
    
    def run(self):
        """Load the file in a separate thread."""
        try:
            if self.file_extension in ['.xlsx', '.xls']:
                content = self._read_excel_file()
            else:
                content = self._read_text_file()
            
            self.file_loaded.emit(content)
            
        except Exception as e:
            self.error_occurred.emit(str(e))
    
    def _read_excel_file(self) -> str:
        """Read Excel file and return content as text."""
        try:
            import openpyxl
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl is required to read Excel files. Please install it with: pip install openpyxl")
        
        try:
            workbook = load_workbook(filename=self.file_path, read_only=True, data_only=True)
            
            worksheet = workbook.active
            if worksheet is None:
                worksheet = workbook.worksheets[0] if workbook.worksheets else None
                if worksheet is None:
                    raise RuntimeError("No worksheets found in the Excel file")
            
            content_lines = []
            row_count = 0
            
            max_row = worksheet.max_row
            progress_interval = max(1, max_row // 100)
            
            for row in worksheet.iter_rows(values_only=True):
                if self.isInterruptionRequested():
                    return ""
                
                row_text = []
                for cell_value in row:
                    if cell_value is not None:
                        row_text.append(str(cell_value))
                
                if row_text:
                    content_lines.append('\t'.join(row_text))
                
                row_count += 1
                
                if row_count % progress_interval == 0:
                    progress = int((row_count / max_row) * 100)
                    self.progress_updated.emit(progress)
            
            workbook.close()
            return '\n'.join(content_lines)
            
        except Exception as e:
            raise RuntimeError(f"Error reading Excel file: {e}")
    
    def _read_text_file(self) -> str:
        """Read text/CSV file and return content."""
        try:
            file_size = os.path.getsize(self.file_path)
            
            if file_size > 1024000:
                content = ""
                with open(self.file_path, 'r', encoding='utf-8-sig') as file:
                    chunk_size = 8192
                    total_chunks = file_size // chunk_size + 1
                    chunk_count = 0
                    
                    while True:
                        if self.isInterruptionRequested():
                            return ""
                        
                        chunk = file.read(chunk_size)
                        if not chunk:
                            break
                        content += chunk
                        chunk_count += 1
                        
                        if chunk_count % 50 == 0:
                            progress = int((chunk_count / total_chunks) * 100)
                            self.progress_updated.emit(progress)
            else:
                with open(self.file_path, 'r', encoding='utf-8-sig') as file:
                    content = file.read()
            
            return content
            
        except UnicodeDecodeError:
            try:
                with open(self.file_path, 'r', encoding='latin-1') as file:
                    content = file.read()
                return content
            except Exception as e:
                raise RuntimeError(f"Could not read file with any encoding: {e}")
        except Exception as e:
            raise RuntimeError(f"Could not read file: {e}")


class ProcessingThread(QThread):
    """Thread for processing large datasets asynchronously to prevent GUI freezing."""
    
    processing_complete = pyqtSignal(list)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)
    
    def __init__(self, extractor, iban_prefix: str, text: str):
        super().__init__()
        self.extractor = extractor
        self.iban_prefix = iban_prefix
        self.text = text
    
    def run(self):
        """Process the text in a separate thread."""
        try:
            lines = self.text.split('\n')
            total_lines = len(lines)
            results = []
            
            for i, line in enumerate(lines):
                if self.isInterruptionRequested():
                    return
                
                line = line.strip()
                if line:
                    phone_numbers = self.extractor.extract_phone_numbers(self.iban_prefix, line)
                    if phone_numbers:
                        results.append({
                            'line_number': i + 1,
                            'text': line,
                            'phone_numbers': phone_numbers,
                            'phone_count': len(phone_numbers)
                        })
                
                if total_lines > 1000 and i % 50 == 0:
                    progress = int((i / total_lines) * 100)
                    self.progress_updated.emit(progress)
                    import time
                    time.sleep(0.001)
            
            self.processing_complete.emit(results)
            
        except Exception as e:
            self.error_occurred.emit(str(e))


class SpanishBankGUI(QMainWindow):
    """Main GUI window for Spanish Bank Phone Extractor."""
    
    def __init__(self):
        super().__init__()
        self.extractor = SpanishBankExtractor()
        self.init_ui()
        self.setup_styles()
    
    def init_ui(self):
        """Initialize the user interface."""
        self.setWindowTitle("Spanish Bank Phone Extractor")
        self.setMinimumSize(800, 600)
        self.resize(1200, 800)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(16)
        
        header = self.create_header()
        main_layout.addWidget(header)
        
        self.tab_widget = QTabWidget()
        self.tab_widget.setFont(QFont("Segoe UI", 10))
        
        self.tab_widget.addTab(self.create_extraction_tab(), "Phone Extraction")
        self.tab_widget.addTab(self.create_bank_info_tab(), "Bank Information")
        
        main_layout.addWidget(self.tab_widget)
    
    def create_header(self) -> QWidget:
        """Create the header section."""
        header_widget = QWidget()
        header_layout = QVBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        title_label = QLabel("Spanish Bank Phone Number Extractor")
        title_label.setFont(QFont("Segoe UI", 24, QFont.Weight.Bold))
        title_label.setProperty("class", "title")
        title_label.setStyleSheet("margin-bottom: 8px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        subtitle_label = QLabel("Extract phone numbers from Spanish bank IBAN data")
        subtitle_label.setFont(QFont("Segoe UI", 12))
        subtitle_label.setProperty("class", "subtitle")
        subtitle_label.setStyleSheet("margin-bottom: 16px;")
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        
        return header_widget
    
    def create_extraction_tab(self) -> QWidget:
        """Create the phone extraction tab."""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setSpacing(20)
        
        bank_section = self.create_bank_selection_section()
        layout.addWidget(bank_section)
        
        input_section = self.create_input_section()
        layout.addWidget(input_section)
        
        self.results_section = self.create_results_section()
        layout.addWidget(self.results_section)
        
        return tab_widget
    
    def create_bank_selection_section(self) -> QWidget:
        """Create the bank selection section."""
        section_widget = QWidget()
        layout = QVBoxLayout(section_widget)
        layout.setSpacing(12)
        
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(12)
        
        self.bank_combo = QComboBox()
        self.bank_combo.setMinimumHeight(40)
        self.bank_combo.setMinimumWidth(300)
        self.bank_combo.setFont(QFont("Segoe UI", 10))
        
        major_banks = self.extractor.get_major_banks()
        self.bank_combo.addItem("Select a bank...", None)
        for iban_prefix, display_name in major_banks:
            self.bank_combo.addItem(display_name, iban_prefix)
        
        self.search_button = ModernButton("Search All Banks", primary=False)
        self.search_button.clicked.connect(self.show_bank_search_dialog)
        
        controls_layout.addWidget(QLabel("Bank:"))
        controls_layout.addWidget(self.bank_combo)
        controls_layout.addWidget(self.search_button)
        controls_layout.addStretch()
        
        self.bank_info_label = QLabel()
        self.bank_info_label.setFont(QFont("Segoe UI", 10))
        self.bank_info_label.setProperty("class", "info")
        self.bank_info_label.setVisible(False)
        
        self.bank_combo.currentIndexChanged.connect(self.on_bank_selected)
        
        layout.addLayout(controls_layout)
        layout.addWidget(self.bank_info_label)
        
        return section_widget
    
    def create_input_section(self) -> QWidget:
        """Create the input section."""
        section_widget = QWidget()
        layout = QVBoxLayout(section_widget)
        layout.setSpacing(12)
        
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(12)
        
        self.file_button = ModernButton("Load File", primary=True)
        self.file_button.clicked.connect(self.load_file)
        
        self.clear_button = ModernButton("Clear", primary=False)
        self.clear_button.clicked.connect(self.clear_input)
        
        self.cancel_button = ModernButton("Cancel", primary=False)
        self.cancel_button.clicked.connect(self.cancel_operation)
        self.cancel_button.setVisible(False)
        
        controls_layout.addWidget(self.file_button)
        controls_layout.addWidget(self.clear_button)
        controls_layout.addWidget(self.cancel_button)
        controls_layout.addStretch()
        
        self.input_text = QTextEdit()
        self.input_text.setFont(QFont("Consolas", 10))
        self.input_text.setStyleSheet("line-height: 1.4; padding: 12px;")
        self.input_text.setPlaceholderText("Paste your data with Spanish IBANs and phone numbers here...")
        
        self.process_button = ModernButton("Extract Phone Numbers", primary=True)
        self.process_button.clicked.connect(self.process_input)
        self.process_button.setEnabled(False)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #404040;
                border-radius: 6px;
                background-color: #2d2d30;
                color: #ffffff;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #0078d4;
                border-radius: 5px;
            }
        """)
        
        layout.addLayout(controls_layout)
        layout.addWidget(self.input_text)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.process_button)
        
        return section_widget
    
    def create_results_section(self) -> QWidget:
        """Create the results section."""
        section_widget = QWidget()
        layout = QVBoxLayout(section_widget)
        layout.setSpacing(12)
        
        # Results table
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(3)
        self.results_table.setHorizontalHeaderLabels(["Line", "Text", "Phone Numbers"])
        header = self.results_table.horizontalHeader()
        if header:
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        
        # Results controls
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(12)
        
        # Export button
        self.export_button = ModernButton("Export Results", primary=False)
        self.export_button.clicked.connect(self.export_results)
        self.export_button.setEnabled(False)
        
        # Clear results button
        self.clear_results_button = ModernButton("Clear Results", primary=False)
        self.clear_results_button.clicked.connect(self.clear_results)
        self.clear_results_button.setEnabled(False)
        
        # Results summary
        self.results_summary = QLabel()
        self.results_summary.setFont(QFont("Segoe UI", 10))
        self.results_summary.setProperty("class", "info")
        
        controls_layout.addWidget(self.export_button)
        controls_layout.addWidget(self.clear_results_button)
        controls_layout.addStretch()
        controls_layout.addWidget(self.results_summary)
        
        layout.addWidget(self.results_table)
        layout.addLayout(controls_layout)
        
        return section_widget
    
    def create_bank_info_tab(self) -> QWidget:
        """Create the bank information tab."""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setSpacing(20)
        
        title_label = QLabel("Spanish Bank Registry")
        title_label.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        title_label.setProperty("class", "title")
        title_label.setStyleSheet("margin-bottom: 16px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        instructions_label = QLabel("Double-click a bank to select it")
        instructions_label.setFont(QFont("Segoe UI", 10))
        instructions_label.setProperty("class", "subtitle")
        instructions_label.setStyleSheet("margin-bottom: 16px; color: #605e5c;")
        instructions_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        search_layout = QHBoxLayout()
        search_layout.setSpacing(12)
        
        self.bank_search_input = QLineEdit()
        self.bank_search_input.setMinimumHeight(40)
        self.bank_search_input.setFont(QFont("Segoe UI", 10))
        self.bank_search_input.setPlaceholderText("Search banks by name...")
        
        self.search_banks_button = ModernButton("Search", primary=True)
        self.search_banks_button.clicked.connect(self.search_banks)
        
        search_layout.addWidget(QLabel("Search:"))
        search_layout.addWidget(self.bank_search_input)
        search_layout.addWidget(self.search_banks_button)
        
        self.banks_table = QTableWidget()
        self.banks_table.setObjectName("banks_table")
        self.banks_table.setColumnCount(4)
        self.banks_table.setHorizontalHeaderLabels(["Entity Code", "Bank Name", "IBAN Prefix", "Address"])
        bank_header = self.banks_table.horizontalHeader()
        if bank_header:
            bank_header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            bank_header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            bank_header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
            bank_header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        
        self.banks_table.itemDoubleClicked.connect(self.on_bank_table_double_clicked)
        
        self.load_all_banks_button = ModernButton("Load All Banks", primary=True)
        self.load_all_banks_button.clicked.connect(self.load_all_banks)
        
        layout.addWidget(title_label)
        layout.addWidget(instructions_label)
        layout.addLayout(search_layout)
        layout.addWidget(self.banks_table)
        layout.addWidget(self.load_all_banks_button)
        
        return tab_widget
    
    def setup_styles(self):
        """Setup application-wide styles."""
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1e1e1e;
                font-family: 'Segoe UI', Arial, sans-serif;
                color: #ffffff;
                min-width: 800px;
                min-height: 600px;
            }
            
            QWidget {
                color: #ffffff;
                background-color: #1e1e1e;
            }
            
            QTabWidget::pane {
                border: 1px solid #404040;
                border-radius: 8px;
                background-color: #2d2d30;
                margin-top: -1px;
            }
            
            QTabBar::tab {
                background-color: #3e3e42;
                border: 1px solid #404040;
                border-bottom: none;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 500;
                color: #cccccc;
            }
            
            QTabBar::tab:selected {
                background-color: #2d2d30;
                border-bottom: 1px solid #2d2d30;
                font-weight: 600;
                color: #ffffff;
            }
            
            QTabBar::tab:hover {
                background-color: #505050;
                color: #ffffff;
            }
            
            QLabel {
                color: #ffffff;
                background-color: transparent;
            }
            
            QTextEdit {
                border: 1px solid #404040;
                border-radius: 6px;
                padding: 8px;
                background-color: #2d2d30;
                color: #ffffff;
                selection-background-color: #0078d4;
                selection-color: white;
                min-height: 100px;
            }
            
            QLineEdit {
                border: 1px solid #404040;
                border-radius: 6px;
                padding: 8px 12px;
                background-color: #2d2d30;
                color: #ffffff;
                selection-background-color: #0078d4;
                selection-color: white;
                min-height: 20px;
            }
            
            QTextEdit:focus, QLineEdit:focus {
                border-color: #0078d4;
                outline: none;
                color: #ffffff;
            }
            
            QComboBox {
                border: 1px solid #404040;
                border-radius: 6px;
                padding: 8px 12px;
                background-color: #2d2d30;
                color: #ffffff;
                min-height: 20px;
            }
            
            QComboBox:focus {
                border-color: #0078d4;
                color: #ffffff;
            }
            
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #cccccc;
                margin-right: 8px;
            }
            
            QComboBox QAbstractItemView {
                border: 1px solid #404040;
                background-color: #2d2d30;
                color: #ffffff;
                selection-background-color: #0078d4;
                selection-color: white;
                border-radius: 4px;
            }
            
            QComboBox QAbstractItemView::item {
                color: #ffffff;
                padding: 8px;
            }
            
            QComboBox QAbstractItemView::item:selected {
                background-color: #0078d4;
                color: white;
            }
            
            QComboBox QAbstractItemView::item:hover {
                background-color: #505050;
                color: #ffffff;
            }
            
            QTableWidget {
                border: 1px solid #404040;
                border-radius: 8px;
                background-color: #2d2d30;
                color: #ffffff;
                gridline-color: #404040;
                alternate-background-color: #252526;
                min-height: 200px;
            }
            
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #404040;
                color: #ffffff;
                background-color: transparent;
            }
            
            QTableWidget::item:selected {
                background-color: #0078d4;
                color: white;
            }
            
            QTableWidget::item:alternate {
                background-color: #252526;
                color: #ffffff;
            }
            
            /* Make bank table rows look clickable */
            QTableWidget#banks_table::item:hover {
                background-color: #3e3e42;
                color: #ffffff;
            }
            
            QTableWidget#banks_table::item:selected {
                background-color: #0078d4;
                color: white;
            }
            
            QHeaderView::section {
                background-color: #3e3e42;
                padding: 12px 8px;
                border: none;
                border-bottom: 1px solid #404040;
                font-weight: 600;
                color: #ffffff;
            }
            
            QScrollBar:vertical {
                background-color: #3e3e42;
                width: 12px;
                border-radius: 6px;
            }
            
            QScrollBar::handle:vertical {
                background-color: #505050;
                border-radius: 6px;
                min-height: 20px;
            }
            
            QScrollBar::handle:vertical:hover {
                background-color: #707070;
            }
            
            QScrollBar:horizontal {
                background-color: #3e3e42;
                height: 12px;
                border-radius: 6px;
            }
            
            QScrollBar::handle:horizontal {
                background-color: #505050;
                border-radius: 6px;
                min-width: 20px;
            }
            
            QScrollBar::handle:horizontal:hover {
                background-color: #707070;
            }
            
            /* Ensure all text elements have proper colors */
            QLabel, QTextEdit, QLineEdit, QComboBox, QTableWidget, QHeaderView {
                color: #ffffff;
            }
            
            /* Specific styling for different text elements */
            QLabel[class="title"] {
                color: #ffffff;
                font-weight: bold;
            }
            
            QLabel[class="subtitle"] {
                color: #cccccc;
            }
            
            QLabel[class="info"] {
                color: #cccccc;
                background-color: #3e3e42;
                border: 1px solid #404040;
                border-radius: 4px;
                padding: 8px;
            }
        """)
    
    def on_bank_selected(self, index: int):
        """Handle bank selection."""
        if index > 0:
            iban_prefix = self.bank_combo.currentData()
            if iban_prefix:
                # Always use normalized prefix for extractor
                normalized_prefix = self.extractor.normalize_iban_prefix(iban_prefix)
                bank_info = self.extractor.get_bank_info(normalized_prefix)
                if bank_info:
                    self.bank_info_label.setText(
                        f"Selected: {bank_info['name']} ({bank_info['entity_code']})"
                    )
                    self.bank_info_label.setVisible(True)
                    self.process_button.setEnabled(True)
                    self.selected_bank_prefix = normalized_prefix  # Store for use in process_input
                else:
                    self.bank_info_label.setVisible(False)
                    self.process_button.setEnabled(False)
                    self.selected_bank_prefix = None
            else:
                self.bank_info_label.setVisible(False)
                self.process_button.setEnabled(False)
                self.selected_bank_prefix = None
        else:
            self.bank_info_label.setVisible(False)
            self.process_button.setEnabled(False)
            self.selected_bank_prefix = None
    
    def show_bank_search_dialog(self):
        """Show bank search dialog."""
        self.tab_widget.setCurrentIndex(1)
        self.load_all_banks()
    
    def load_file(self):
        """Load file into input text area."""
        file_name, _ = QFileDialog.getOpenFileName(
            self, 
            "Select File", 
            "", 
            "Text Files (*.txt);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        
        if file_name:
            # Start async file loading immediately
            self._start_async_file_loading(file_name)
    
    def _start_async_file_loading(self, file_name: str):
        """Start asynchronous file loading with proper UI feedback."""
        try:
            # Check file size before loading (this is quick, so it's OK to do in main thread)
            file_size = os.path.getsize(file_name)
            if file_size > 10485760:  # 10MB
                reply = QMessageBox.question(
                    self, "Large File", 
                    f"File size is {file_size / 1024 / 1024:.1f}MB. This may take a while to load. Continue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return
            
            # Show loading cursor and progress bar
            self.setCursor(Qt.CursorShape.WaitCursor)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.progress_bar.setFormat("Loading file... %p%")
            
            # Disable buttons during loading
            self.process_button.setEnabled(False)
            self.file_button.setEnabled(False)
            self.cancel_button.setVisible(True) # Show cancel button
            
            # Determine file type and read accordingly
            file_extension = os.path.splitext(file_name)[1].lower()
            
            # Create and start the file loader thread
            self.file_loader = FileLoaderThread(file_name, file_extension)
            self.file_loader.file_loaded.connect(self.on_file_loaded)
            self.file_loader.error_occurred.connect(self.on_file_error)
            self.file_loader.progress_updated.connect(self.update_progress_bar)
            self.file_loader.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not read file: {e}")
            self.progress_bar.setVisible(False)
            self.process_button.setEnabled(False)
            self.file_button.setEnabled(True)
            self.cancel_button.setVisible(False)
    
    def on_file_loaded(self, content: str):
        """Handle file loading completion."""
        self.input_text.setPlainText(content)
        
        # Restore UI state
        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.progress_bar.setVisible(False)
        self.process_button.setEnabled(True)
        self.file_button.setEnabled(True)
        self.cancel_button.setVisible(False)
        
        # Clear previous results
        self.export_button.setEnabled(False)  # No results yet
        self.clear_results_button.setEnabled(False)
        self.results_summary.clear()
        self.results_table.setRowCount(0)
        
        # Show success message for large files
        if len(content) > 100000:  # More than 100KB
            QMessageBox.information(self, "Success", "File loaded successfully!")
    
    def on_file_error(self, error_message: str):
        """Handle file loading errors."""
        QMessageBox.critical(self, "Error", f"Failed to load file: {error_message}")
        
        # Restore UI state
        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.progress_bar.setVisible(False)
        self.process_button.setEnabled(False)
        self.file_button.setEnabled(True)
        self.cancel_button.setVisible(False)
        self.export_button.setEnabled(False)
        self.clear_results_button.setEnabled(False)
        self.results_summary.clear()
        self.results_table.setRowCount(0)
    
    def update_progress_bar(self, progress: int):
        """Update the progress bar during file loading."""
        if hasattr(self, 'progress_bar') and self.progress_bar.isVisible():
            # Only update if progress actually changed to reduce UI overhead
            if self.progress_bar.value() != progress:
                self.progress_bar.setValue(progress)
                # Process events less frequently to maintain responsiveness
                if progress % 5 == 0:  # Only process events every 5% progress
                    QApplication.processEvents()
    
    def clear_input(self):
        """Clear input text area."""
        self.input_text.clear()
    
    def process_input(self):
        """Process input text to extract phone numbers."""
        iban_prefix = getattr(self, 'selected_bank_prefix', None)
        if not iban_prefix:
            QMessageBox.warning(self, "Warning", "Please select a bank first.")
            return
        text = self.input_text.toPlainText()
        if not text.strip():
            QMessageBox.warning(self, "Warning", "Please enter some text to process.")
            return
        
        # Check if this is a large dataset that needs async processing
        if len(text) > 50000:  # More than 50KB
            # Use async processing for large datasets
            self._process_large_dataset(iban_prefix, text)
        else:
            # Use synchronous processing for smaller datasets
            self._process_small_dataset(iban_prefix, text)
    
    def _process_large_dataset(self, iban_prefix: str, text: str):
        """Process large datasets asynchronously."""
        # Show processing cursor and progress bar
        self.setCursor(Qt.CursorShape.WaitCursor)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Processing data... %p%")
        
        # Disable buttons during processing
        self.process_button.setEnabled(False)
        self.file_button.setEnabled(False)
        self.cancel_button.setVisible(True) # Show cancel button
        
        # Create and start the processing thread
        self.processing_thread = ProcessingThread(self.extractor, iban_prefix, text)
        self.processing_thread.processing_complete.connect(self.on_processing_complete)
        self.processing_thread.error_occurred.connect(self.on_processing_error)
        self.processing_thread.progress_updated.connect(self.update_progress_bar)
        self.processing_thread.start()
    
    def _process_small_dataset(self, iban_prefix: str, text: str):
        """Process small datasets synchronously."""
        try:
            # Show processing cursor for small files
            self.setCursor(Qt.CursorShape.WaitCursor)
            
            # Process the text using the extractor
            results = self.extractor.process_text(iban_prefix, text)
            self.display_results(results)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error processing data: {e}")
        finally:
            # Restore cursor
            self.setCursor(Qt.CursorShape.ArrowCursor)
    
    def on_processing_complete(self, results):
        """Handle processing completion."""
        self.display_results(results)
        
        # Restore UI state
        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.progress_bar.setVisible(False)
        self.process_button.setEnabled(True)
        self.file_button.setEnabled(True)
        self.cancel_button.setVisible(False)
        
        # Show success message for large datasets
        if len(results) > 100:
            QMessageBox.information(self, "Success", f"Processing complete! Found {len(results)} lines with phone numbers.")
    
    def on_processing_error(self, error_message: str):
        """Handle processing errors."""
        QMessageBox.critical(self, "Error", f"Error processing data: {error_message}")
        
        # Restore UI state
        self.setCursor(Qt.CursorShape.ArrowCursor)
        self.progress_bar.setVisible(False)
        self.process_button.setEnabled(True)
        self.file_button.setEnabled(True)
        self.cancel_button.setVisible(False)
    
    def display_results(self, results):
        """Display results in the table."""
        # Block signals for better performance
        self.results_table.blockSignals(True)
        
        self.results_table.setRowCount(len(results))
        
        total_phones = 0
        for i, result in enumerate(results):
            # Line number
            line_item = QTableWidgetItem(str(result['line_number']))
            line_item.setFlags(line_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.results_table.setItem(i, 0, line_item)
            
            # Text
            text_item = QTableWidgetItem(result['text'])
            text_item.setFlags(text_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.results_table.setItem(i, 1, text_item)
            
            # Phone numbers
            phones_text = ', '.join(result['phone_numbers'])
            phones_item = QTableWidgetItem(phones_text)
            phones_item.setFlags(phones_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.results_table.setItem(i, 2, phones_item)
            
            total_phones += len(result['phone_numbers'])
        
        # Unblock signals
        self.results_table.blockSignals(False)
        
        # Update summary
        self.results_summary.setText(f"Found {len(results)} lines with {total_phones} phone numbers")
        
        # Enable buttons
        self.export_button.setEnabled(len(results) > 0)
        self.clear_results_button.setEnabled(len(results) > 0)
    
    def export_results(self):
        """Export results to a file."""
        file_name, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Results", 
            "extracted_phones.txt",
            "Text Files (*.txt);;CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        
        if file_name:
            try:
                # Show saving cursor
                self.setCursor(Qt.CursorShape.WaitCursor)
                
                # Determine file type and export accordingly
                file_extension = os.path.splitext(file_name)[1].lower()
                
                if file_extension == '.xlsx':
                    self._export_to_excel(file_name)
                else:
                    self._export_to_text(file_name)
                
                QMessageBox.information(self, "Success", f"Phone numbers exported to {file_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Could not export results: {e}")
            finally:
                # Restore cursor
                self.setCursor(Qt.CursorShape.ArrowCursor)
    
    def _export_to_excel(self, file_path: str):
        """Export results to Excel file."""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is required to export Excel files. Please install it with: pip install openpyxl")
        
        # Create a new workbook
        workbook = Workbook()
        worksheet = workbook.active
        if worksheet is None:
            # Create a new worksheet if none exists
            worksheet = workbook.create_sheet("Extracted Phone Numbers")
        else:
            worksheet.title = "Extracted Phone Numbers"
        
        # Add headers
        worksheet['A1'] = "Line Number"
        worksheet['B1'] = "Original Text"
        worksheet['C1'] = "Phone Numbers"
        
        # Add data
        row_num = 2
        for table_row in range(self.results_table.rowCount()):
            line_item = self.results_table.item(table_row, 0)
            text_item = self.results_table.item(table_row, 1)
            phone_item = self.results_table.item(table_row, 2)
            
            if line_item:
                worksheet[f'A{row_num}'] = line_item.text()
            if text_item:
                worksheet[f'B{row_num}'] = text_item.text()
            if phone_item:
                worksheet[f'C{row_num}'] = phone_item.text()
            
            row_num += 1
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(file_path)
        workbook.close()
    
    def _export_to_text(self, file_path: str):
        """Export results to text/CSV file."""
        with open(file_path, 'w', encoding='utf-8') as file:
            for row in range(self.results_table.rowCount()):
                phone_item = self.results_table.item(row, 2)
                if phone_item:
                    phones = phone_item.text()
                    # Split phone numbers and write each one on a separate line
                    phone_list = [phone.strip() for phone in phones.split(',')]
                    for phone in phone_list:
                        if phone:  # Only write non-empty phone numbers
                            file.write(f"{phone}\n")
    
    def clear_results(self):
        """Clear results table."""
        self.results_table.setRowCount(0)
        self.results_summary.clear()
        self.export_button.setEnabled(False)
        self.clear_results_button.setEnabled(False)
    
    def search_banks(self):
        """Search banks by name."""
        search_term = self.bank_search_input.text().strip()
        if not search_term:
            self.load_all_banks()
            return
        
        matches = self.extractor.search_banks(search_term)
        self.display_banks(matches)
    
    def load_all_banks(self):
        """Load all banks into the table."""
        all_banks = self.extractor.get_all_banks()
        self.display_banks(all_banks)
    
    def on_bank_table_double_clicked(self, item):
        """Handle double-click on bank table item."""
        row = item.row()
        iban_prefix_item = self.banks_table.item(row, 2)  # IBAN Prefix column
        if iban_prefix_item:
            iban_prefix = iban_prefix_item.text()
            self.select_bank_from_table(iban_prefix)
    
    def select_bank_from_table(self, iban_prefix):
        """Select a bank from the table and switch to extraction tab."""
        # Find the bank in the dropdown and select it
        found = False
        for i in range(self.bank_combo.count()):
            if self.bank_combo.itemData(i) == iban_prefix:
                self.bank_combo.setCurrentIndex(i)
                found = True
                break
        
        # If found in dropdown, trigger the selection event manually
        if found:
            # Manually trigger the bank selection logic
            self.on_bank_selected(self.bank_combo.currentIndex())
        
        # Always manually set the bank selection state to ensure it works
        normalized_prefix = self.extractor.normalize_iban_prefix(iban_prefix)
        bank_info = self.extractor.get_bank_info(normalized_prefix)
        if bank_info:
            self.selected_bank_prefix = normalized_prefix
            self.bank_info_label.setText(
                f"Selected: {bank_info['name']} ({bank_info['entity_code']})"
            )
            self.bank_info_label.setVisible(True)
            self.process_button.setEnabled(True)
        
        # Switch to the extraction tab
        self.tab_widget.setCurrentIndex(0)
    
    def display_banks(self, banks):
        """Display banks in the table."""
        self.banks_table.setRowCount(len(banks))
        
        for i, bank_data in enumerate(banks):
            if len(bank_data) == 4:
                iban_prefix, name, entity_code, address = bank_data
            else:
                iban_prefix, name = bank_data
                entity_code = iban_prefix[2:] if len(iban_prefix) >= 6 else iban_prefix
                bank_info = self.extractor.get_bank_info(iban_prefix)
                address = bank_info['address'] if bank_info else ""
            
            # Entity code
            entity_item = QTableWidgetItem(entity_code)
            entity_item.setFlags(entity_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.banks_table.setItem(i, 0, entity_item)
            
            # Bank name
            name_item = QTableWidgetItem(name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.banks_table.setItem(i, 1, name_item)
            
            # IBAN prefix
            prefix_item = QTableWidgetItem(iban_prefix)
            prefix_item.setFlags(prefix_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.banks_table.setItem(i, 2, prefix_item)
            
            # Address
            address_item = QTableWidgetItem(address)
            address_item.setFlags(address_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.banks_table.setItem(i, 3, address_item)

    def cancel_operation(self):
        """Cancel the current operation (file loading or processing)."""
        reply = QMessageBox.question(
            self, "Confirm Cancel",
            "Are you sure you want to cancel the current operation? This will stop the file loading or processing.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            # Request interruption for graceful termination
            if hasattr(self, 'file_loader') and self.file_loader.isRunning():
                self.file_loader.requestInterruption()
                self.file_loader.wait(5000)  # Wait for thread to finish
            
            if hasattr(self, 'processing_thread') and self.processing_thread.isRunning():
                self.processing_thread.requestInterruption()
                self.processing_thread.wait(5000)  # Wait for thread to finish
            
            # Restore UI state
            self.setCursor(Qt.CursorShape.ArrowCursor)
            self.progress_bar.setVisible(False)
            self.process_button.setEnabled(True)
            self.file_button.setEnabled(True)
            self.cancel_button.setVisible(False)
            QMessageBox.information(self, "Operation Cancelled", "Operation cancelled.")


def main():
    """Main function to run the GUI application."""
    app = QApplication(sys.argv)
    app.setApplicationName("Spanish Bank Phone Extractor")
    app.setApplicationVersion("1.0")
    
    # Set application style
    app.setStyle("Fusion")
    
    # Ensure proper color scheme handling
    palette = app.palette()
    palette.setColor(QPalette.ColorRole.Window, QColor("#1e1e1e"))
    palette.setColor(QPalette.ColorRole.WindowText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Base, QColor("#2d2d30"))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#252526"))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#2d2d30"))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Text, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Button, QColor("#3e3e42"))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.BrightText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Link, QColor("#0078d4"))
    palette.setColor(QPalette.ColorRole.Highlight, QColor("#0078d4"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("white"))
    app.setPalette(palette)
    
    # Create and show the main window
    window = SpanishBankGUI()
    window.show()
    
    # Run the application
    sys.exit(app.exec())


if __name__ == "__main__":
    main() 